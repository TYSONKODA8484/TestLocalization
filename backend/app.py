from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import json
import csv
import io
import os
import sys
import time
import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openai import OpenAI
from dotenv import load_dotenv

# Ensure the backend directory is in the path for imports
backend_dir = os.path.dirname(os.path.abspath(__file__))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

# Load environment variables from .env file in backend directory
load_dotenv(os.path.join(backend_dir, '.env'))

# Import prompt functions
from prompts import get_prompt_for_scenario, get_qa_prompt

app = Flask(__name__)
CORS(app)

# Initialize OpenAI client
client = None
def get_openai_client():
    global client
    if client is None:
        api_key = os.getenv('OPENAI_API_KEY')
        if not api_key or api_key == 'your_openai_api_key_here':
            return None
        client = OpenAI(api_key=api_key)
    return client

@app.route('/api/health', methods=['GET'])
def health():
    has_api_key = bool(os.getenv('OPENAI_API_KEY') and os.getenv('OPENAI_API_KEY') != 'your_openai_api_key_here')
    return jsonify({
        'status': 'ok',
        'openai_configured': has_api_key
    })

# ==================== TRANSLATION ENDPOINTS ====================

@app.route('/api/translate', methods=['POST'])
def translate():
    """
    Translate a single text string to a target language.
    Uses scenario-specific prompts for optimal results.
    """
    try:
        openai_client = get_openai_client()
        if not openai_client:
            return jsonify({'error': 'OpenAI API key not configured'}), 500
        
        data = request.json
        text = data.get('text', '')
        target_language = data.get('targetLanguage', '')
        scenario = data.get('scenario', 'general')
        location = data.get('location', '')
        
        if not text or not target_language:
            return jsonify({'error': 'Missing text or targetLanguage'}), 400
        
        # Get scenario-specific prompt
        system_prompt, user_prompt = get_prompt_for_scenario(scenario, text, target_language, location)
        
        # Call OpenAI
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=500
        )
        
        translation = response.choices[0].message.content.strip()
        
        return jsonify({
            'translation': translation,
            'source': text,
            'targetLanguage': target_language,
            'scenario': scenario
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/translate/batch', methods=['POST'])
def translate_batch():
    """
    Translate multiple texts to multiple languages.
    Returns translations progressively for each text-language pair.
    """
    try:
        openai_client = get_openai_client()
        if not openai_client:
            return jsonify({'error': 'OpenAI API key not configured'}), 500
        
        data = request.json
        texts = data.get('texts', [])
        languages = data.get('languages', [])
        scenario = data.get('scenario', 'general')
        location = data.get('location', '')
        
        if not texts or not languages:
            return jsonify({'error': 'Missing texts or languages'}), 400
        
        results = []
        
        for text in texts:
            text_result = {
                'source': text,
                'translations': {}
            }
            
            for lang in languages:
                try:
                    # Get scenario-specific prompt
                    system_prompt, user_prompt = get_prompt_for_scenario(scenario, text, lang, location)
                    
                    # Call OpenAI with rate limiting
                    response = openai_client.chat.completions.create(
                        model="gpt-4o",
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_prompt}
                        ],
                        temperature=0.1,
                        max_tokens=500
                    )
                    
                    translation = response.choices[0].message.content.strip()
                    text_result['translations'][lang] = translation
                    
                    # Small delay to avoid rate limiting
                    time.sleep(0.1)
                    
                except Exception as e:
                    text_result['translations'][lang] = f'[Error: {str(e)}]'
            
            results.append(text_result)
        
        return jsonify({'results': results})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/verify', methods=['POST'])
def verify_translations():
    """
    Verify and correct translations for the provided table.

    Request JSON:
    {
      tableData: [{ source, translations: { langName: translation } }],
      languages: ["Spanish", ...],
      scenario: "general",
      chunkSize: 50
    }

    Response JSON:
    {
      results: [{ source, translations: { langName: corrected } }]
    }
    """
    try:
        openai_client = get_openai_client()
        if not openai_client:
            return jsonify({'error': 'OpenAI API key not configured'}), 500

        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        scenario = data.get('scenario', 'general')
        chunk_size = int(data.get('chunkSize', 50))

        # Prepare output structure
        corrected = [
            {
                'source': row.get('source', ''),
                'translations': dict(row.get('translations', {}))
            }
            for row in table_data
        ]

        qa_issues = []

        # For each language, QA in chunks
        for lang in languages:
            # Build list of entries for this language
            entries = [
                {
                    'source': row.get('source', ''),
                    'translation': row.get('translations', {}).get(lang, '')
                }
                for row in table_data
            ]

            # Chunk processing to avoid token limits
            for start in range(0, len(entries), chunk_size):
                chunk = entries[start:start+chunk_size]
                system_prompt, user_prompt = get_qa_prompt(chunk, lang, scenario)

                response = openai_client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    temperature=0.1,
                    max_tokens=1500
                )

                content = response.choices[0].message.content
                # Parse JSON safely
                try:
                    import json
                    corrected_list = json.loads(content)
                    # Apply corrections back to the main table
                    for idx, item in enumerate(corrected_list):
                        global_index = start + idx
                        if 0 <= global_index < len(corrected):
                            original_value = corrected[global_index]['translations'].get(lang, '')
                            new_value = item.get('translation', original_value)
                            corrected[global_index]['translations'][lang] = new_value

                            notes = item.get('notes', []) or []
                            if new_value != original_value or notes:
                                qa_issues.append({
                                    'source': corrected[global_index].get('source', ''),
                                    'language': lang,
                                    'original': original_value,
                                    'corrected': new_value,
                                    'notes': notes
                                })
                except Exception:
                    # Leave as-is if parsing fails, annotate error in translation
                    for idx in range(start, min(start+chunk_size, len(entries))):
                        corrected[idx]['translations'][lang] = corrected[idx]['translations'].get(lang, '')

                # Small delay between chunks
                time.sleep(0.1)

        return jsonify({'results': corrected, 'issues': qa_issues})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/translate/stream', methods=['POST'])
def translate_stream():
    """
    Stream translations one at a time for progressive UI updates.
    Client should call this for each text-language pair.
    """
    try:
        openai_client = get_openai_client()
        if not openai_client:
            return jsonify({'error': 'OpenAI API key not configured'}), 500
        
        data = request.json
        text = data.get('text', '')
        target_language = data.get('targetLanguage', '')
        scenario = data.get('scenario', 'general')
        location = data.get('location', '')
        row_index = data.get('rowIndex', 0)
        
        if not text or not target_language:
            return jsonify({'error': 'Missing text or targetLanguage'}), 400
        
        # Get scenario-specific prompt
        system_prompt, user_prompt = get_prompt_for_scenario(scenario, text, target_language, location)
        
        # Call OpenAI
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=500
        )
        
        translation = response.choices[0].message.content.strip()
        
        return jsonify({
            'translation': translation,
            'source': text,
            'targetLanguage': target_language,
            'rowIndex': row_index,
            'scenario': scenario
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'rowIndex': data.get('rowIndex', 0)}), 500

# ==================== EXPORT ENDPOINTS ====================

@app.route('/api/export/csv', methods=['POST'])
def export_csv():
    """Export table data as CSV (tab-separated for Excel compatibility)"""
    try:
        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter='\t')
        
        # Header
        writer.writerow(['Source'] + languages)
        
        # Data rows
        for row in table_data:
            row_data = [row.get('source', '')]
            for lang in languages:
                row_data.append(row.get('translations', {}).get(lang, ''))
            writer.writerow(row_data)
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='localization.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export table data as Excel with formatting"""
    try:
        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Localization"
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        headers = ['Source'] + languages
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, row in enumerate(table_data, 2):
            ws.cell(row=row_idx, column=1, value=row.get('source', '')).border = border
            for col_idx, lang in enumerate(languages, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get('translations', {}).get(lang, ''))
                cell.border = border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='localization.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/json', methods=['POST'])
def export_json():
    """Export table data as JSON"""
    try:
        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        
        result = []
        for row in table_data:
            item = {'source': row.get('source', '')}
            for lang in languages:
                item[lang] = row.get('translations', {}).get(lang, '')
            result.append(item)
        
        output = json.dumps(result, indent=2, ensure_ascii=False)
        
        return send_file(
            io.BytesIO(output.encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name='localization.json'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/xml', methods=['POST'])
def export_xml():
    """Export table data as XML"""
    try:
        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        
        root = ET.Element('localization')
        
        for row in table_data:
            entry = ET.SubElement(root, 'entry')
            source = ET.SubElement(entry, 'source')
            source.text = row.get('source', '')
            
            translations = ET.SubElement(entry, 'translations')
            for lang in languages:
                lang_elem = ET.SubElement(translations, 'language')
                lang_elem.set('name', lang)
                lang_elem.text = row.get('translations', {}).get(lang, '')
        
        # Pretty print
        xml_str = minidom.parseString(ET.tostring(root, encoding='unicode')).toprettyxml(indent='  ')
        
        return send_file(
            io.BytesIO(xml_str.encode('utf-8')),
            mimetype='application/xml',
            as_attachment=True,
            download_name='localization.xml'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/ios', methods=['POST'])
def export_ios():
    """Export as iOS .strings file (one per language)"""
    try:
        data = request.json
        table_data = data.get('tableData', [])
        language = data.get('language', 'English')
        
        lines = ['/* iOS Localizable.strings */']
        lines.append(f'/* Language: {language} */')
        lines.append('')
        
        for row in table_data:
            source = row.get('source', '').replace('"', '\\"')
            translation = row.get('translations', {}).get(language, source).replace('"', '\\"')
            # Use source as key
            key = source.replace(' ', '_').replace('"', '').lower()
            lines.append(f'"{key}" = "{translation}";')
        
        output = '\n'.join(lines)
        
        return send_file(
            io.BytesIO(output.encode('utf-8')),
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'Localizable_{language}.strings'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/android', methods=['POST'])
def export_android():
    """Export as Android strings.xml file (one per language)"""
    try:
        data = request.json
        table_data = data.get('tableData', [])
        language = data.get('language', 'English')
        
        root = ET.Element('resources')
        root.set('xmlns:tools', 'http://schemas.android.com/tools')
        
        # Add comment
        comment = ET.Comment(f' {language} strings.xml ')
        root.append(comment)
        
        for row in table_data:
            source = row.get('source', '')
            translation = row.get('translations', {}).get(language, source)
            # Use source as key (snake_case)
            key = source.replace(' ', '_').replace('"', '').lower()
            
            string_elem = ET.SubElement(root, 'string')
            string_elem.set('name', key)
            string_elem.text = translation
        
        # Pretty print
        xml_str = '<?xml version="1.0" encoding="utf-8"?>\n'
        xml_str += minidom.parseString(ET.tostring(root, encoding='unicode')).toprettyxml(indent='    ')
        # Remove extra declaration
        xml_str = xml_str.replace('<?xml version="1.0" ?>\n', '')
        
        return send_file(
            io.BytesIO(xml_str.encode('utf-8')),
            mimetype='application/xml',
            as_attachment=True,
            download_name=f'strings_{language}.xml'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/ios-all', methods=['POST'])
def export_ios_all():
    """Export iOS .strings files for all languages as zip"""
    try:
        import zipfile
        
        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for language in languages:
                lines = ['/* iOS Localizable.strings */']
                lines.append(f'/* Language: {language} */')
                lines.append('')
                
                for row in table_data:
                    source = row.get('source', '').replace('"', '\\"')
                    translation = row.get('translations', {}).get(language, source).replace('"', '\\"')
                    key = source.replace(' ', '_').replace('"', '').lower()
                    lines.append(f'"{key}" = "{translation}";')
                
                output = '\n'.join(lines)
                zip_file.writestr(f'{language}/Localizable.strings', output.encode('utf-8'))
        
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name='ios_strings.zip'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/android-all', methods=['POST'])
def export_android_all():
    """Export Android strings.xml files for all languages as zip"""
    try:
        import zipfile
        
        data = request.json
        table_data = data.get('tableData', [])
        languages = data.get('languages', [])
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for language in languages:
                root = ET.Element('resources')
                
                for row in table_data:
                    source = row.get('source', '')
                    translation = row.get('translations', {}).get(language, source)
                    key = source.replace(' ', '_').replace('"', '').lower()
                    
                    string_elem = ET.SubElement(root, 'string')
                    string_elem.set('name', key)
                    string_elem.text = translation
                
                xml_str = '<?xml version="1.0" encoding="utf-8"?>\n'
                xml_str += minidom.parseString(ET.tostring(root, encoding='unicode')).toprettyxml(indent='    ')
                xml_str = xml_str.replace('<?xml version="1.0" ?>\n', '')
                
                # Android folder convention
                folder = f'values-{language[:2].lower()}' if language != 'English' else 'values'
                zip_file.writestr(f'{folder}/strings.xml', xml_str.encode('utf-8'))
        
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name='android_strings.zip'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
